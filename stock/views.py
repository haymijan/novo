# stock/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from .models import Warehouse, Location
from .forms import WarehouseForm, LocationForm
from .models import LotSerialNumber
from .forms import LotSerialNumberForm
from .forms import InventoryTransactionForm
from django.db import transaction
from django.core.exceptions import ValidationError
from django.http import JsonResponse
from .forms import DateRangeForm
from datetime import timedelta
from django.db.models import Q
from django.core.paginator import Paginator
from .forms import StockMovementFilterForm
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from .forms import TransactionFilterForm
from django.db.models import Sum
from products.models import Product
from .forms import LotFilterForm

from django.contrib import messages
from .models import InventoryTransaction, Stock, Location, Warehouse, LotSerialNumber
from .forms import LotBasedInventoryAdjustmentForm # <-- নতুন فرم ইম্পোর্ট করা হয়েছে
from .services import StockService


# --- Warehouse CRUD Views ---
@login_required
@permission_required('stock.view_warehouse', login_url='/admin/')
def warehouse_list(request):
    warehouses = Warehouse.objects.all().order_by('name')
    context = {'warehouses': warehouses, 'title': 'All Warehouses'}
    return render(request, 'stock/warehouse_list.html', context)

@login_required
@permission_required('stock.add_warehouse', login_url='/admin/')
def add_warehouse(request):
    if request.method == 'POST':
        form = WarehouseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock:warehouse_list')
    else:
        form = WarehouseForm()
    return render(request, 'stock/add_warehouse.html', {'form': form, 'title': 'Add New Warehouse'})

@login_required
@permission_required('stock.change_warehouse', login_url='/admin/')
def edit_warehouse(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        form = WarehouseForm(request.POST, instance=warehouse)
        if form.is_valid():
            form.save()
            return redirect('stock:warehouse_list')
    else:
        form = WarehouseForm(instance=warehouse)
    return render(request, 'stock/edit_warehouse.html', {'form': form, 'title': f'Edit Warehouse: {warehouse.name}'})

@login_required
@permission_required('stock.delete_warehouse', login_url='/admin/')
def delete_warehouse(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        warehouse.delete()
        return redirect('stock:warehouse_list')
    return render(request, 'confirm_delete.html', {'object': warehouse, 'title': f'Confirm Delete Warehouse: {warehouse.name}'})

# --- Location CRUD Views ---
@login_required
@permission_required('stock.view_location', login_url='/admin/')
def location_list(request):
    locations = Location.objects.all().order_by('name')
    context = {'locations': locations, 'title': 'All Locations'}
    return render(request, 'stock/location_list.html', context)

@login_required
@permission_required('stock.add_location', login_url='/admin/')
def add_location(request):
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock:location_list')
    else:
        form = LocationForm()
    return render(request, 'stock/add_location.html', {'form': form, 'title': 'Add New Location'})

@login_required
@permission_required('stock.change_location', login_url='/admin/')
def edit_location(request, pk):
    location = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        form = LocationForm(request.POST, instance=location)
        if form.is_valid():
            form.save()
            return redirect('stock:location_list')
    else:
        form = LocationForm(instance=location)
    return render(request, 'stock/edit_location.html', {'form': form, 'title': f'Edit Location: {location.name}'})

@login_required
@permission_required('stock.delete_location', login_url='/admin/')
def delete_location(request, pk):
    location = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        location.delete()
        return redirect('stock:location_list')
    return render(request, 'confirm_delete.html', {'object': location, 'title': f'Confirm Delete Location: {location.name}'})

@login_required
@permission_required('stock.view_lotserialnumber', login_url='/admin/')
def lot_serial_list(request):
    user = request.user
    lots_queryset = LotSerialNumber.objects.select_related(
        'product', 'location__warehouse'
    ).filter(quantity__gt=0)

    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            lots_queryset = lots_queryset.filter(location__warehouse=user_warehouse)
        else:
            lots_queryset = LotSerialNumber.objects.none()

    form = LotFilterForm(request.GET, user=user)
    if form.is_valid():
        product = form.cleaned_data.get('product')
        warehouse = form.cleaned_data.get('warehouse')

        if product:
            lots_queryset = lots_queryset.filter(product=product)
        if warehouse and user.is_superuser:
            lots_queryset = lots_queryset.filter(location__warehouse=warehouse)

    lots = lots_queryset.order_by('product__name', 'location__name')

    context = {
        'lots': lots,
        'title': 'Available Lot/Serial Numbers',
        'form': form
    }
    return render(request, 'stock/lot_serial_list.html', context)

@login_required
@permission_required('stock.add_lotserialnumber', login_url='/admin/')
def add_lot_serial(request):
    if request.method == 'POST':
        form = LotSerialNumberForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock:lot_serial_list')
    else:
        form = LotSerialNumberForm()
    context = {'form': form, 'title': 'Add New Lot/Serial Number'}
    return render(request, 'stock/add_lot_serial.html', context)

@login_required
@permission_required('stock.change_lotserialnumber', login_url='/admin/')
def edit_lot_serial(request, pk):
    lot = get_object_or_404(LotSerialNumber, pk=pk)
    if request.method == 'POST':
        form = LotSerialNumberForm(request.POST, instance=lot)
        if form.is_valid():
            form.save()
            return redirect('stock:lot_serial_list')
    else:
        form = LotSerialNumberForm(instance=lot)
    context = {'form': form, 'title': f'Edit Lot/Serial: {lot.lot_number}'}
    return render(request, 'stock/edit_lot_serial.html', context)

@login_required
@permission_required('stock.delete_lotserialnumber', login_url='/admin/')
def delete_lot_serial(request, pk):
    lot = get_object_or_404(LotSerialNumber, pk=pk)
    if request.method == 'POST':
        lot.delete()
        return redirect('stock:lot_serial_list')
    return render(request, 'confirm_delete.html', {'object': lot, 'title': f'Confirm Delete Lot/Serial: {lot.lot_number}'})

@login_required
def transaction_list(request):
    transactions_queryset = InventoryTransaction.objects.select_related(
        'product', 'warehouse', 'source_location__warehouse', 'destination_location__warehouse', 'lot_serial', 'user'
    ).order_by('-transaction_date')

    user = request.user
    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            transactions_queryset = transactions_queryset.filter(
                Q(source_location__warehouse=user_warehouse) | 
                Q(destination_location__warehouse=user_warehouse) |
                Q(warehouse=user_warehouse)
            ).distinct()

    form = TransactionFilterForm(request.GET, user=request.user)
    if form.is_valid():
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        selected_user = form.cleaned_data.get('user')
        warehouse = form.cleaned_data.get('warehouse')

        if start_date:
            transactions_queryset = transactions_queryset.filter(transaction_date__date__gte=start_date)
        if end_date:
            transactions_queryset = transactions_queryset.filter(transaction_date__date__lte=end_date)
        if selected_user:
            transactions_queryset = transactions_queryset.filter(user=selected_user)
        if warehouse and user.is_superuser:
            transactions_queryset = transactions_queryset.filter(
                Q(source_location__warehouse=warehouse) | 
                Q(destination_location__warehouse=warehouse) |
                Q(warehouse=warehouse)
            ).distinct()

    paginator = Paginator(transactions_queryset, 20)
    page_number = request.GET.get('page')
    transactions = paginator.get_page(page_number)

    context = {
        'title': 'All Transactions',
        'transactions': transactions,
        'form': form,
    }
    return render(request, 'stock/transaction_list.html', context)

@login_required
@permission_required('stock.add_inventorytransaction', login_url='/admin/')
def record_transaction(request):
    if request.method == 'POST':
        form = InventoryTransactionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock:transaction_list')
    else:
        form = InventoryTransactionForm()
    return render(request, 'stock/record_transaction.html', {'form': form, 'title': 'Record Inventory Transaction'})

@login_required
@permission_required('stock.add_inventorytransaction', login_url='/admin/')
def inventory_adjustment(request):
    if request.method == 'POST':
        form = LotBasedInventoryAdjustmentForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                with transaction.atomic():
                    lot_to_adjust = form.cleaned_data['lot']
                    new_quantity = form.cleaned_data['new_quantity']
                    StockService.adjust_stock_for_lot(
                        lot_serial=lot_to_adjust,
                        new_quantity=new_quantity,
                        user=request.user,
                        notes=f"Manual inventory adjustment by user {request.user.username}."
                    )
                
                messages.success(request, f"Stock for Lot #{lot_to_adjust.lot_number} has been successfully adjusted.")
                return redirect('stock:lot_serial_list')

            except Exception as e:
                messages.error(request, f"An error occurred: {e}")
    else:
        form = LotBasedInventoryAdjustmentForm(user=request.user)
        
    context = {
        'form': form, 
        'title': 'Inventory Adjustment (Lot Based)'
    }
    return render(request, 'stock/inventory_adjustment.html', context)

@login_required
def get_products_by_warehouse_ajax(request):
    warehouse_id = request.GET.get('warehouse_id')
    if not warehouse_id:
        return JsonResponse({'products': []})

    products = Product.objects.filter(
        stocks__warehouse_id=warehouse_id,
        stocks__quantity__gt=0
    ).distinct().order_by('name')
    
    product_data = [{'id': p.id, 'text': p.name} for p in products]
    return JsonResponse({'products': product_data})

@login_required
def get_lots_by_product_ajax(request):
    product_id = request.GET.get('product_id')
    warehouse_id = request.GET.get('warehouse_id')
    
    if not product_id or not warehouse_id:
        return JsonResponse({'lots': []})
        
    lots = LotSerialNumber.objects.filter(
        product_id=product_id,
        location__warehouse_id=warehouse_id,
        quantity__gt=0
    ).order_by('created_at')
    
    lots_data = [
        {
            'id': lot.id, 
            'text': f"{lot.lot_number} (Current Qty: {int(lot.quantity)})"
        } for lot in lots
    ]
    return JsonResponse({'lots': lots_data})

@login_required
@permission_required('stock.change_inventorytransaction', login_url='/admin/')
def edit_transaction(request, pk):
    transaction_obj = get_object_or_404(InventoryTransaction, pk=pk)
    if request.method == 'POST':
        form = InventoryTransactionForm(request.POST, instance=transaction_obj)
        if form.is_valid():
            form.save()
            return redirect('stock:transaction_list')
    else:
        form = InventoryTransactionForm(instance=transaction_obj)
    return render(request, 'stock/edit_transaction.html', {'form': form, 'title': f'Edit Transaction: {transaction_obj.product.name}'})

@login_required
@permission_required('stock.delete_inventorytransaction', login_url='/admin/')
def delete_transaction(request, pk):
    transaction_obj = get_object_or_404(InventoryTransaction, pk=pk)
    if request.method == 'POST':
        transaction_obj.delete()
        return redirect('stock:transaction_list')
    return render(request, 'confirm_delete.html', {'object': transaction_obj, 'title': 'Confirm Delete Transaction'})

@login_required
def check_product_tracking(request, product_id):
    try:
        product = Product.objects.get(pk=product_id)
        data = {'tracking_method': product.tracking_method}
        if product.tracking_method in ['lot', 'serial']:
            lots = LotSerialNumber.objects.filter(product=product, quantity__gt=0).values('id', 'lot_number', 'quantity')
            data['lots'] = list(lots)
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    
@login_required
def stock_movement_report(request):

    transactions_list = InventoryTransaction.objects.select_related(
        'product', 'source_location__warehouse', 'destination_location__warehouse'
    ).order_by('-transaction_date')

    user = request.user
    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            transactions_list = transactions_list.filter(
                Q(source_location__warehouse=user_warehouse) |
                Q(destination_location__warehouse=user_warehouse)
            )

    form = StockMovementFilterForm(request.GET)
    if form.is_valid():
        product = form.cleaned_data.get('product')
        transaction_type = form.cleaned_data.get('transaction_type')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        warehouse = form.cleaned_data.get('warehouse')

        if product:
            transactions_list = transactions_list.filter(product=product)
        if transaction_type:
            transactions_list = transactions_list.filter(transaction_type=transaction_type)
        if start_date:
            transactions_list = transactions_list.filter(transaction_date__date__gte=start_date)
        if end_date:
            transactions_list = transactions_list.filter(transaction_date__date__lte=end_date)

        if warehouse:
            if user.is_superuser:
                 transactions_list = transactions_list.filter(
                    Q(source_location__warehouse=warehouse) |
                    Q(destination_location__warehouse=warehouse)
                )

    paginator = Paginator(transactions_list, 20)
    page_number = request.GET.get('page')
    transactions = paginator.get_page(page_number)

    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            form.fields['warehouse'].queryset = Warehouse.objects.filter(pk=user_warehouse.pk)
            form.fields['warehouse'].initial = user_warehouse
            form.fields['warehouse'].widget.attrs['disabled'] = True


    context = {
        'title': 'Stock Movement Report',
        'transactions': transactions,
        'form': form,
    }
    return render(request, 'stock/stock_movement_report.html', context)

@login_required
def get_available_lots(request):
    product_id = request.GET.get('product_id')
    location_id = request.GET.get('location_id')
    
    if not product_id or not location_id:
        return JsonResponse([], safe=False)

    lots = LotSerialNumber.objects.filter(
        product_id=product_id, 
        location_id=location_id,
        quantity__gt=0
    ).values('id', 'lot_number', 'quantity')
    
    return JsonResponse(list(lots), safe=False)

def get_lots_by_location_and_product(request):
    product_id = request.GET.get('product_id')
    location_id = request.GET.get('location_id')
    
    lots = LotSerialNumber.objects.filter(
        product_id=product_id,
        location_id=location_id,
        quantity__gt=0
    ).values('id', 'lot_number', 'quantity')
    
    return JsonResponse(list(lots), safe=False)

@login_required
def download_transaction_report(request):
    transactions_queryset = InventoryTransaction.objects.select_related(
        'product', 'user', 'source_location__warehouse', 'destination_location__warehouse'
    ).order_by('-transaction_date')

    user = request.user
    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            transactions_queryset = transactions_queryset.filter(
                Q(source_location__warehouse=user_warehouse) | Q(destination_location__warehouse=user_warehouse)
            )

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_user_id = request.GET.get('user')
    warehouse_id = request.GET.get('warehouse')

    if start_date:
        transactions_queryset = transactions_queryset.filter(transaction_date__date__gte=start_date)
    if end_date:
        transactions_queryset = transactions_queryset.filter(transaction_date__date__lte=end_date)
    if selected_user_id:
        transactions_queryset = transactions_queryset.filter(user_id=selected_user_id)
    if warehouse_id and user.is_superuser:
        transactions_queryset = transactions_queryset.filter(
            Q(source_location__warehouse_id=warehouse_id) | Q(destination_location__warehouse_id=warehouse_id)
        )

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transaction_report.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Transactions"

    headers = ['Date', 'Product', 'Type', 'Quantity', 'User', 'Source', 'Destination', 'Notes']
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for transaction in transactions_queryset:
        worksheet.append([
            transaction.transaction_date.strftime('%Y-%m-%d %H:%M'),
            transaction.product.name,
            transaction.get_transaction_type_display(),
            transaction.quantity,
            transaction.user.username if transaction.user else 'N/A',
            f"{transaction.source_location.name} ({transaction.source_location.warehouse.name})" if transaction.source_location else 'N/A',
            f"{transaction.destination_location.name} ({transaction.destination_location.warehouse.name})" if transaction.destination_location else 'N/A',
            transaction.notes or ''
        ])

    for col_num, column_title in enumerate(headers, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 20

    workbook.save(response)
    return response

@login_required
def product_stock_details(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    user = request.user
    stock_by_warehouse = Stock.objects.filter(product=product, quantity__gt=0).values(
        'warehouse__name'
    ).annotate(
        total_stock=Sum('quantity')
    ).order_by('warehouse__name')

    lot_details = None
    user_warehouse = getattr(user, 'warehouse', None)
    
    if user.is_superuser:

        lot_details = LotSerialNumber.objects.filter(
            product=product, 
            quantity__gt=0
        ).select_related('location', 'location__warehouse').order_by('location__warehouse__name', 'expiration_date')
    
    elif user.groups.filter(name='Branch Manager').exists() and user_warehouse:

        lot_details = LotSerialNumber.objects.filter(
            product=product, 
            location__warehouse=user_warehouse,
            quantity__gt=0
        ).select_related('location').order_by('expiration_date')

    context = {
        'title': f'Stock Details for {product.name}',
        'product': product,
        'stock_by_warehouse': stock_by_warehouse,
        'lot_details': lot_details,
        'user_warehouse': user_warehouse
    }
    return render(request, 'stock/product_stock_details.html', context)