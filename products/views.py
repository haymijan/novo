# products/views.py

import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Sum, Count, F, Q
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import timedelta
from django.db.models.functions import TruncMonth # For monthly trend charts
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import IntegrityError

# Local Application Imports
from .models import Product, Category, UnitOfMeasureCategory, UnitOfMeasure, Brand
from .forms import ProductForm, CategoryForm, UnitOfMeasureCategoryForm, UnitOfMeasureForm, BrandForm
from sales.models import SalesOrder
from partners.models import Supplier, Customer 
from products.models import Product
from stock.models import Stock
from stock.models import Warehouse

# Standard Library Imports
from io import BytesIO
import os

from .admin import ProductResource
from .forms import ProductImportForm
from tablib import Dataset

# Third-Party Imports for Excel Export
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image as PillowImage

# Third-Party Imports for PDF Export
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl.utils import get_column_letter

import datetime

#======================================================================

DEFAULT_CURRENCY_SYMBOL = 'QAR '

def apply_product_filters(request, base_queryset):
    query = request.GET.get('q')
    category_id = request.GET.get('category')
    brand_id = request.GET.get('brand')
    status = request.GET.get('status')
    warehouse_id = request.GET.get('warehouse')

    filtered_queryset = base_queryset
    
    if query:
        filtered_queryset = filtered_queryset.filter(Q(name__icontains=query) | Q(product_code__icontains=query))

    if category_id:
        try:
            category = Category.objects.get(id=category_id)
            descendant_ids = category.get_all_children_ids()
            filtered_queryset = filtered_queryset.filter(category__id__in=descendant_ids)
        except Category.DoesNotExist:
            pass
            
    if brand_id:
        filtered_queryset = filtered_queryset.filter(brand_id=brand_id)
    if warehouse_id:
        filtered_queryset = filtered_queryset.filter(stocks__warehouse_id=warehouse_id)

    if status:
        if status == 'in_stock':
            filtered_queryset = filtered_queryset.filter(calculated_total_quantity__gt=F('min_stock_level'))
        elif status == 'low_stock':
            filtered_queryset = filtered_queryset.filter(
                calculated_total_quantity__lte=F('min_stock_level'),
                calculated_total_quantity__gt=0
            )
        elif status == 'out_of_stock':
            filtered_queryset = filtered_queryset.filter(calculated_total_quantity__lte=0)
    return filtered_queryset.distinct()


@login_required
@permission_required('products.view_product', login_url='/admin/')
def product_list(request):
    user = request.user
    user_warehouse = getattr(user, 'warehouse', None)


    products_query = Product.objects.select_related('category', 'brand')
    
    quantity_annotation_filter = Q()
    if not user.is_superuser and user_warehouse:
        quantity_annotation_filter = Q(stocks__warehouse=user_warehouse)
    
    products_query = products_query.annotate(
        calculated_total_quantity=Coalesce(Sum('stocks__quantity', filter=quantity_annotation_filter), 0)
    ).order_by('name')

    filtered_products = apply_product_filters(request, products_query)

    paginator = Paginator(filtered_products, 15)
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)
    
    context = {
        'products': products_page,
        'title': 'All Products',
        'all_categories': Category.objects.all().order_by('name'),
        'all_brands': Brand.objects.all().order_by('name'),
        'product_statuses': Product.STATUS_CHOICES,
        'all_warehouses': Warehouse.objects.all().order_by('name'),
        'DEFAULT_CURRENCY_SYMBOL': 'QAR ',
    }

    if request.headers.get('HX-Request'):
        return render(request, 'products/partials/product_list_content.html', context)
    
    return render(request, 'products/product_list.html', context)

@login_required
@permission_required('products.add_product', login_url='/admin/')
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('products:product_list')
    else:
        form = ProductForm()
    return render(request, 'products/add_product.html', {'form': form, 'title': 'Add New Product'})

@login_required
@permission_required('products.change_product', login_url='/admin/')
def edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('products:product_list')
        else:
            print("Form Errors:", form.errors)
            print("Non Field Errors:", form.non_field_errors())
    else:
        form = ProductForm(instance=product)
    return render(request, 'products/edit_product.html', {'form': form, 'title': f'Edit Product: {product.name}'})

@login_required
@permission_required('products.delete_product', login_url='/admin/')
def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('products:product_list')
    return render(request, 'confirm_delete.html', {'object': product, 'title': f'Confirm Delete: {product.name}'})

# category_list ভিউটি আপডেট করুন
@login_required
@permission_required('products.view_category', login_url='/admin/')
def category_list(request):
    categories = Category.objects.select_related('parent').all().order_by('name')
    return render(request, 'products/category_list.html', {'categories': categories, 'title': 'All Categories'})

@login_required
@permission_required('products.add_category', login_url='/admin/')
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('products:category_list')
    else:
        form = CategoryForm()
    return render(request, 'products/add_category.html', {'form': form, 'title': 'Add New Category'})

@login_required
@permission_required('products.change_category', login_url='/admin/')
def edit_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('products:category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'products/edit_category.html', {'form': form, 'title': f'Edit Category: {category.name}'})

@login_required
@permission_required('products.delete_category', login_url='/admin/')
def delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        return redirect('products:category_list')
    return render(request, 'confirm_delete.html', {'object': category, 'title': 'Confirm Delete Category'})

# --- Unit of Measure (UoM) Category Views ---
@login_required
@permission_required('products.view_unitofmeasurecategory', login_url='/admin/')
def uom_category_list(request):
    uom_categories = UnitOfMeasureCategory.objects.all().order_by('name')
    return render(request, 'products/uom_category_list.html', {'uom_categories': uom_categories, 'title': 'Unit of Measure Categories'})

@login_required
@permission_required('products.add_unitofmeasurecategory', login_url='/admin/')
def add_uom_category(request):
    if request.method == 'POST':
        form = UnitOfMeasureCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('products:uom_category_list')
    else:
        form = UnitOfMeasureCategoryForm()
    return render(request, 'products/add_uom_category.html', {'form': form, 'title': 'Add New UoM Category'})

@login_required
@permission_required('products.change_unitofmeasurecategory', login_url='/admin/')
def edit_uom_category(request, pk):
    uom_category = get_object_or_404(UnitOfMeasureCategory, pk=pk)
    if request.method == 'POST':
        form = UnitOfMeasureCategoryForm(request.POST, instance=uom_category)
        if form.is_valid():
            form.save()
            return redirect('products:uom_category_list')
    else:
        form = UnitOfMeasureCategoryForm(instance=uom_category)
    return render(request, 'products/edit_uom_category.html', {'form': form, 'title': f'Edit UoM Category: {uom_category.name}'})

@login_required
@permission_required('products.delete_unitofmeasurecategory', login_url='/admin/')
def delete_uom_category(request, pk):
    uom_category = get_object_or_404(UnitOfMeasureCategory, pk=pk)
    if request.method == 'POST':
        uom_category.delete()
        return redirect('products:uom_category_list')
    return render(request, 'confirm_delete.html', {'object': uom_category, 'title': 'Confirm Delete UoM Category'})

# --- Unit of Measure (UoM) Views ---
@login_required
@permission_required('products.view_unitofmeasure', login_url='/admin/')
def unit_of_measure_list(request):
    units_of_measure = UnitOfMeasure.objects.all().order_by('name')
    context = {'units_of_measure': units_of_measure, 'title': 'Units of Measure'}
    return render(request, 'products/unit_of_measure_list.html', context)

@login_required
@permission_required('products.add_unitofmeasure', login_url='/admin/')
def add_unit_of_measure(request):
    if request.method == 'POST':
        form = UnitOfMeasureForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('products:unit_of_measure_list')
    else:
        form = UnitOfMeasureForm()
    return render(request, 'products/add_unit_of_measure.html', {'form': form, 'title': 'Add New Unit of Measure'})

@login_required
@permission_required('products.change_unitofmeasure', login_url='/admin/')
def edit_unit_of_measure(request, pk):
    unit_of_measure = get_object_or_404(UnitOfMeasure, pk=pk)
    if request.method == 'POST':
        form = UnitOfMeasureForm(request.POST, instance=unit_of_measure)
        if form.is_valid():
            form.save()
            return redirect('products:unit_of_measure_list')
    else:
        form = UnitOfMeasureForm(instance=unit_of_measure)
    return render(request, 'products/edit_unit_of_measure.html', {'form': form, 'title': f'Edit Unit of Measure: {unit_of_measure.name}'})

@login_required
@permission_required('products.delete_unitofmeasure', login_url='/admin/')
def delete_unit_of_measure(request, pk):
    unit_of_measure = get_object_or_404(UnitOfMeasure, pk=pk)
    if request.method == 'POST':
        unit_of_measure.delete()
        return redirect('products:unit_of_measure_list')
    return render(request, 'confirm_delete.html', {'object': unit_of_measure, 'title': 'Confirm Delete Unit of Measure'})

# --- AJAX Views ---
def get_product_sale_price(request):
    product_id = request.GET.get('product_id')
    product = get_object_or_404(Product, id=product_id)
    return JsonResponse({'sale_price': product.sale_price})

@login_required
@permission_required('products.view_product', login_url='/admin/')
def export_products_excel(request):
    user = request.user
    user_warehouse = getattr(user, 'warehouse', None)

    selected_warehouse_id = request.GET.get('warehouse')

    # --- ১. ডাটা ফিল্টারিং লজিক ---
    quantity_annotation_filter = Q()

    if selected_warehouse_id:
        quantity_annotation_filter = Q(stocks__warehouse_id=selected_warehouse_id)
    elif not user.is_superuser and user_warehouse:
        quantity_annotation_filter = Q(stocks__warehouse=user_warehouse)

    products_query = Product.objects.select_related('category', 'brand').annotate(
        calculated_total_quantity=Coalesce(Sum('stocks__quantity', filter=quantity_annotation_filter), 0)
    ).order_by('name')

    filtered_products = apply_product_filters(request, products_query)

    # ব্রাঞ্চ ইনফো
    report_branch_info = "All Branches"
    warehouse_id = request.GET.get('warehouse')
    if user.is_superuser and warehouse_id:
        try:
            selected_warehouse = Warehouse.objects.get(id=warehouse_id)
            report_branch_info = f"Branch: {selected_warehouse.name}"
        except Warehouse.DoesNotExist:
            pass 
    elif not user.is_superuser and user_warehouse:
        report_branch_info = f"Branch: {user_warehouse.name}"

    # --- ২. এক্সেল ফাইল তৈরি ও ডিজাইন ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    # --- পেজ সেটআপ (A4 & Printing) ---
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE # ল্যান্ডস্কেপ মোড (বেশি কলামের জন্য ভালো)
    ws.page_setup.fitToWidth = 1  # এক পেজের প্রস্থে ফিট করবে
    ws.page_setup.fitToHeight = 0 # উচ্চতা অটোমেটিক বাড়ব
    ws.print_title_rows = '4:4'   # ৪ নম্বর রো (হেডার) প্রতি পেজে প্রিন্ট হবে

    # --- স্টাইল ডেফিনিশন ---
    # বর্ডার স্টাইল
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    # হেডার স্টাইল (গাঢ় নীল ব্যাকগ্রাউন্ড, সাদা টেক্সট)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # টাইটেল স্টাইল
    title_font = Font(bold=True, size=16, color="1F4E78")
    subtitle_font = Font(size=10, italic=True, color="555555")

    # --- ৩. রিপোর্ট হেডার সেকশন ---
    ws.merge_cells('A1:I1')
    ws['A1'] = "SHOP ERP - INVENTORY STATUS REPORT"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = f"{report_branch_info} | Generated on: {datetime.date.today().strftime('%d-%b-%Y')}"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.append([]) # একটি ফাঁকা রো (Row 3)

    # --- ৪. টেবিল হেডার ---
    headers = ['Product Name', 'SKU/Code', 'Category', 'Brand', 'Purchase Price', 'Cost Price', 'Sale Price', 'Stock Qty', 'Status']
    ws.append(headers) # Row 4

    # হেডার ডিজাইন অ্যাপ্লাই
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # --- ৫. ডাটা পপুলেশন ---
    row_num = 5
    for product in filtered_products:
        # স্ট্যাটাস লজিক
        if not product.is_active:
            status_text = "Inactive"
        elif product.calculated_total_quantity > product.min_stock_level:
            status_text = "In Stock"
        elif product.calculated_total_quantity > 0:
            status_text = "Low Stock"
        else:
            status_text = "Out of Stock"

        row_data = [
            product.name,
            product.product_code,
            product.category.name if product.category else '-',
            product.brand.name if product.brand else '-',
            product.price,
            product.cost_price,
            product.sale_price,
            product.calculated_total_quantity,
            status_text,
        ]
        ws.append(row_data)

        # প্রতিটি সেলে স্টাইল অ্যাপ্লাই করা
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            
            # অ্যালাইনমেন্ট এবং কালার
            if col_num in [5, 6, 7, 8]: # প্রাইস এবং স্টক কলাম
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00' if col_num != 8 else '0'
            else:
                cell.alignment = Alignment(horizontal='left')

            # জেব্রা স্ট্রাইপ (Zebra Striping) - জোড় সংখ্যার রো-তে হালকা রঙ
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # স্ট্যাটাস কলাম কালারিং (Optional)
            if col_num == 9: # Status Column
                cell.alignment = Alignment(horizontal='center')
                if status_text == "Out of Stock":
                    cell.font = Font(color="FF0000", bold=True)
                elif status_text == "Low Stock":
                    cell.font = Font(color="FFA500", bold=True)
                elif status_text == "In Stock":
                    cell.font = Font(color="008000", bold=True)

        row_num += 1

    # --- ৬. কলাম প্রস্থ অটো-ফিট (Merged Cell এরর ফিক্স সহ) ---
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        
        for cell in col:
            try:
                # হেডার বা মার্জ করা সেল স্কিপ করা হচ্ছে যাতে কলাম বিশাল না হয়ে যায়
                if cell.row <= 3: 
                    continue
                
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # মিনিমাম এবং ম্যাক্সিমাম উইডথ সেট করা
        adjusted_width = (max_length + 2)
        if adjusted_width < 10: adjusted_width = 10
        if adjusted_width > 40: adjusted_width = 40
        
        ws.column_dimensions[column_letter].width = adjusted_width

    # রেসপন্স রিটার্ন
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Inventory_Report_A4.xlsx"'
    wb.save(response)
    
    return response


@login_required
@permission_required('products.view_product', login_url='/admin/')
def export_products_pdf(request):
    user = request.user
    user_warehouse = getattr(user, 'warehouse', None)

    selected_warehouse_id = request.GET.get('warehouse')

    quantity_annotation_filter = Q()

    if selected_warehouse_id:
        quantity_annotation_filter = Q(stocks__warehouse_id=selected_warehouse_id)
    elif not user.is_superuser and user_warehouse:
        quantity_annotation_filter = Q(stocks__warehouse=user_warehouse)

    products_query = Product.objects.select_related('category', 'brand').annotate(
        calculated_total_quantity=Coalesce(Sum('stocks__quantity', filter=quantity_annotation_filter), 0)
    ).order_by('name')

    filtered_products = apply_product_filters(request, products_query)
    
    report_branch_info = "All Branches"
    warehouse_id = request.GET.get('warehouse')

    if user.is_superuser and warehouse_id:
        try:
            selected_warehouse = Warehouse.objects.get(id=warehouse_id)
            report_branch_info = f"Branch: {selected_warehouse.name}"
        except Warehouse.DoesNotExist:
            pass 
    elif not user.is_superuser and user_warehouse:
        report_branch_info = f"Branch: {user_warehouse.name}"

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    story = []

    styles.add(ParagraphStyle(name='ReportHeading', fontSize=18, alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=5))
    styles.add(ParagraphStyle(name='SubHeading', fontSize=10, alignment=TA_CENTER, spaceAfter=20))
    styles.add(ParagraphStyle(name='TableHeader', fontSize=8, alignment=TA_CENTER, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='TableCell', fontSize=7, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='TableCellLeft', fontSize=7, alignment=TA_LEFT))

    story.append(Paragraph("Product Inventory Report", styles['ReportHeading']))
    story.append(Paragraph(f"({report_branch_info})", styles['SubHeading']))
    
    table_data = [
        [Paragraph(h, styles['TableHeader']) for h in ['Product Name', 'SKU', 'Category', 'Brand', 'Cost Price', 'Sale Price', 'Qty']]
    ]
    
    for product in filtered_products:
        table_data.append([
            Paragraph(product.name, styles['TableCellLeft']),
            Paragraph(product.product_code or 'N/A', styles['TableCell']),
            Paragraph(product.category.name if product.category else 'N/A', styles['TableCell']),
            Paragraph(product.brand.name if product.brand else 'N/A', styles['TableCell']),
            Paragraph(f"{product.cost_price:.2f}", styles['TableCell']),
            Paragraph(f"{product.sale_price:.2f}", styles['TableCell']),
            Paragraph(str(product.calculated_total_quantity), styles['TableCell']),
        ])

    table = Table(table_data, colWidths=[2.5*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        # --- এই লাইনটিতে বানান ঠিক করা হয়েছে ---
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke), 
        # --- whitesmokoke এর পরিবর্তে whitesmoke ---
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="products_inventory.pdf"'
    return response

@login_required
def print_product_labels(request):
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')

        if not product_ids:
            return redirect('products:product_list')

        labels_to_print = []
        for pid in product_ids:
            try:
                product = Product.objects.get(id=int(pid))
                quantity = int(request.POST.get(f'quantity_{pid}', 1))
                
                if product.barcode and hasattr(product.barcode, 'path'):
                    for _ in range(quantity):
                        labels_to_print.append({
                            'name': product.name,
                            'price': f"{product.sale_price:.2f}",
                            'currency': DEFAULT_CURRENCY_SYMBOL,
                            'barcode_path': product.barcode.path,
                            'product_code': product.product_code
                        })
            except (Product.DoesNotExist, ValueError):
                continue

        if not labels_to_print:
            return redirect('products:product_list')

        # --- Thermal Printer-এর জন্য PDF তৈরির প্রক্রিয়া ---
        buffer = BytesIO()
        label_width, label_height = 2.25 * inch, 1.25 * inch
        doc = SimpleDocTemplate(buffer, pagesize=(label_width, label_height), leftMargin=0.05*inch, rightMargin=0.05*inch, topMargin=0.05*inch, bottomMargin=0.05*inch)
        story = []
        
        for label_data in labels_to_print:
            product_name = label_data['name']
            if len(product_name) > 30:
                 name_style = ParagraphStyle('LabelName', fontSize=6, alignment=TA_CENTER, leading=7)
            else:
                 name_style = ParagraphStyle('LabelName', fontSize=7, alignment=TA_CENTER, leading=8)

            inner_table_data = [
                [Paragraph(product_name, name_style)],
                [RLImage(label_data['barcode_path'], width=1.8*inch, height=0.4*inch)],
                [
                    Paragraph(label_data['product_code'], ParagraphStyle('LabelCode', fontName='Helvetica', fontSize=7, alignment=TA_LEFT)),
                    Paragraph(f"<b>{label_data['currency']}{label_data['price']}</b>", ParagraphStyle('LabelPrice', fontName='Helvetica-Bold', fontSize=10, alignment=TA_RIGHT))
                ]
            ]
            
            inner_table = Table(inner_table_data, 
                                colWidths=[1.025*inch, 1.025*inch], 
                                rowHeights=[0.25*inch, 0.45*inch, 0.2*inch])
            
            inner_table.setStyle(TableStyle([
                ('SPAN', (0,0), (1,0)), ('SPAN', (0,1), (1,1)),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,2), (0,2), 'LEFT'), ('ALIGN', (1,2), (1,2), 'RIGHT'),
                ('VALIGN', (0,2), (1,2), 'BOTTOM'),
            ]))
            
            story.append(inner_table)

        doc.build(story)
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

    return redirect('products:product_list')

@login_required
@permission_required('products.change_unitofmeasurecategory', login_url='/admin/')
def edit_uom_category(request, pk):
    uom_category = get_object_or_404(UnitOfMeasureCategory, pk=pk)
    if request.method == 'POST':
        form = UnitOfMeasureCategoryForm(request.POST, instance=uom_category)
        if form.is_valid():
            form.save()
            return redirect('products:uom_category_list')
    else:
        form = UnitOfMeasureCategoryForm(instance=uom_category)
    return render(request, 'products/edit_uom_category.html', {'form': form, 'title': f'Edit UoM Category: {uom_category.name}'})

@login_required
@permission_required('products.delete_unitofmeasurecategory', login_url='/admin/')
def delete_uom_category(request, pk):
    uom_category = get_object_or_404(UnitOfMeasureCategory, pk=pk)
    if request.method == 'POST':
        uom_category.delete()
        return redirect('products:uom_category_list')
    return render(request, 'confirm_delete.html', {'object': uom_category, 'title': 'Confirm Delete UoM Category'})

@login_required
@permission_required('stock.view_stock', login_url='/admin/')
def product_stock_by_location(request, pk):
    product = get_object_or_404(Product, pk=pk)
    stocks = Stock.objects.filter(product=product).select_related('location').order_by('location__name')
    
    total_stock = sum(s.quantity for s in stocks)
    
    context = {
        'title': f'Stock for {product.name}',
        'product': product,
        'stocks': stocks,
        'total_stock': total_stock,
    }
    return render(request, 'products/product_stock_by_location.html', context)

@login_required
@permission_required('products.view_brand', login_url='/admin/')
def brand_list(request):
    brands = Brand.objects.all().order_by('name')
    return render(request, 'products/brand_list.html', {'brands': brands, 'title': 'All Brands'})

@login_required
@permission_required('products.add_brand', login_url='/admin/')
def add_brand(request):
    if request.method == 'POST':
        form = BrandForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Brand added successfully!')
            return redirect('products:brand_list')
    else:
        form = BrandForm()
    return render(request, 'products/add_brand.html', {'form': form, 'title': 'Add New Brand'})

@login_required
@permission_required('products.change_brand', login_url='/admin/')
def edit_brand(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        form = BrandForm(request.POST, instance=brand)
        if form.is_valid():
            form.save()
            messages.success(request, f'Brand "{brand.name}" updated successfully!')
            return redirect('products:brand_list')
    else:
        form = BrandForm(instance=brand)
    return render(request, 'products/edit_brand.html', {'form': form, 'title': f'Edit Brand: {brand.name}'})

@login_required
@permission_required('products.delete_brand', login_url='/admin/')
def delete_brand(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        brand.delete()
        messages.success(request, f'Brand "{brand.name}" deleted successfully!')
        return redirect('products:brand_list')
    return render(request, 'confirm_delete.html', {'object': brand, 'title': 'Confirm Delete Brand'})

@login_required
@permission_required('products.change_product', login_url='/admin/')
def product_bulk_action(request):
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')
        action = request.POST.get('action')

        if not product_ids:
            messages.error(request, "Please select at least one product.")
            return redirect('products:product_list')

        if action == 'activate':
            Product.objects.filter(id__in=product_ids).update(is_active=True)
            messages.success(request, f"{len(product_ids)} product(s) have been activated.")
        elif action == 'deactivate':
            Product.objects.filter(id__in=product_ids).update(is_active=False)
            messages.success(request, f"{len(product_ids)} product(s) have been deactivated.")
        else:
            messages.warning(request, "Invalid action selected.")

    return redirect('products:product_list')

@login_required
@permission_required('products.add_product', login_url='/admin/')
def import_products_view(request):
    if request.method == 'POST':
        form = ProductImportForm(request.POST, request.FILES)
        if form.is_valid():
            product_resource = ProductResource()
            dataset = Dataset()
            new_products_file = request.FILES['file']

            try:
                if new_products_file.name.endswith('csv'):
                    dataset.load(new_products_file.read().decode('utf-8'), format='csv')
                elif new_products_file.name.endswith(('xls', 'xlsx')):
                    dataset.load(new_products_file.read(), format='xlsx')
                else:
                    messages.error(request, "Unsupported file format. Please upload a CSV or Excel file.")
                    return redirect('products:import_products')

                # 'dry_run' ব্যবহার করে প্রথমে ডেটা যাচাই করা হচ্ছে
                result = product_resource.import_data(dataset, dry_run=True, raise_errors=True)

                # কোনো ভুল না থাকলে ডেটা সেভ করা হচ্ছে
                product_resource.import_data(dataset, dry_run=False)
                messages.success(request, "Products imported successfully!")
                return redirect('products:product_list')

            except IntegrityError as e:
                messages.error(request, f"Data integrity error: {e}")
            except Exception as e:
                # ফাইলের ভেতরের ভুলের জন্য
                error_html = "<ul>"
                for error in getattr(e, 'errors', []):
                     error_html += f"<li>Row {error.row}: {', '.join(error.error.messages)}</li>"
                if not getattr(e, 'errors', []):
                    error_html += f"<li>An unexpected error occurred: {e}</li>"
                error_html += "</ul>"
                messages.error(request, f"There were errors in the file:{error_html}")


    else:
        form = ProductImportForm()

    context = {
        'form': form,
        'title': 'Import Products'
    }
    return render(request, 'products/import_products.html', context)

@login_required
def download_sample_template(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Product Import Template"

    # সঠিক কলাম হেডার (মডেল অনুযায়ী)
    headers = [
        'id', 'name', 'product_code', 'category', 'brand', 
        'price', 'cost_price', 'sale_price', 
        'min_stock_level', 'tracking_method', 'is_active', 'description'
    ]
    sheet.append(headers)
    
    # হেডার স্টাইল করা (বোল্ড এবং গ্রে ব্যাকগ্রাউন্ড)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4E73DF", end_color="4E73DF", fill_type="solid")
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # --- স্মার্ট স্যাম্পল ডাটা (Hierarchy ও SKU সহ) ---
    sample_data = [
        ['', 'Samsung Galaxy S24', 'SAM-S24-001', 'Mobile', 'Samsung', 80000, 75000, 85000, 5, 'serial', 1, 'Latest flagship phone'],
        ['', 'Logitech Mouse', 'LOG-M-102', 'Accessories', 'Logitech', 800, 600, 1000, 20, 'none', 1, 'Wireless mouse'],
        ['', 'HP Laptop Charger', 'HP-CH-01', 'Laptop Parts', 'HP', 1500, 1200, 1800, 10, 'none', 1, 'Original adapter'],
        # আপডেটের উদাহরণ (ID সহ)
        ['101', 'Old Product Update', 'UPD-001', 'General', 'Generic', 500, 400, 600, 5, 'none', 1, 'Updated description example'],
    ]

    for row_data in sample_data:
        sheet.append(row_data)

    # কলামের প্রস্থ অটোমেটিক ঠিক করা
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="product_import_template_v2.xlsx"'
    
    workbook.save(response)
    return response