# partners/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse
from django.contrib import messages
from decimal import Decimal
from django.db.models import Q, Sum, F, Count, Max, Value, DecimalField
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

# --- Model Imports (Corrected) ---

from .models import Customer, Supplier
from .forms import CustomerForm, SupplierForm, CustomerFilterForm

# Purchase App
from purchase.models import PurchaseReturn, SupplierCreditNote

# Finance App (AP & AR) -> এই লাইনগুলো ঠিক করা হয়েছে
from finance.ap.models import VendorBill, BillPayment
from finance.ar.models import CustomerInvoice, InvoicePayment # [FIXED]: সঠিক লোকেশন থেকে ইম্পোর্ট
from finance.gl.models import CustomerRefund

# Sales App -> এই লাইনটি ঠিক করা হয়েছে
from sales.models import SalesOrder, SalesReturn, SalesReturnItem # [FIXED]: এখানে আর CustomerInvoice নেই

# POS & Marketing
from pos.models import POSOrder
from marketing.models import GiftCard


#=============================================================================
# VIEWS
#=============================================================================

@login_required
@permission_required('partners.view_supplier', login_url='/admin/')
def supplier_list(request):
    suppliers = Supplier.objects.all().order_by('name')
    return render(request, 'partners/supplier_list.html', {'suppliers': suppliers, 'title': 'All Suppliers'})

@login_required
@permission_required('partners.add_supplier', login_url='/admin/')
def add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Supplier added successfully!')
            return redirect('partners:supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'partners/add_supplier.html', {'form': form, 'title': 'Add New Supplier'})

@login_required
@permission_required('partners.change_supplier', login_url='/admin/')
def edit_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, f'Supplier "{supplier.name}" updated successfully!')
            return redirect('partners:supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'partners/edit_supplier.html', {'form': form, 'title': f'Edit Supplier: {supplier.name}'})

@login_required
@permission_required('partners.delete_supplier', login_url='/admin/')
def delete_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier_name = supplier.name
        supplier.delete()
        messages.success(request, f'Supplier "{supplier_name}" deleted successfully!')
        return redirect('partners:supplier_list')
    return render(request, 'confirm_delete.html', {'object': supplier, 'title': f'Confirm Delete: {supplier.name}'})


@login_required
@permission_required('partners.view_customer', login_url='/admin/')
def customer_list(request):
    customers_list = Customer.objects.all().order_by('-created_at')

    filter_form = CustomerFilterForm(request.GET or None)
    if filter_form.is_valid():
        query = filter_form.cleaned_data.get('q')
        start_date = filter_form.cleaned_data.get('start_date')
        end_date = filter_form.cleaned_data.get('end_date')

        if query:
            customers_list = customers_list.filter(
                Q(name__icontains=query) |
                Q(email__icontains=query) |
                Q(phone__icontains=query)
            )
        if start_date:
            customers_list = customers_list.filter(created_at__date__gte=start_date)
        if end_date:
            customers_list = customers_list.filter(created_at__date__lte=end_date)

    # Annotations for listing page
    customers_list = customers_list.annotate(
        total_sales_amount=Sum(
            'salesorder__total_amount', 
            filter=Q(salesorder__status__in=['delivered', 'partially_delivered']),
            distinct=True
        ),
        total_received_amount=Sum(
            F('salesorder__paid_by_cash') + F('salesorder__paid_by_card'),
            filter=Q(salesorder__status__in=['delivered', 'partially_delivered']),
            distinct=True
        ),
        # Note: Return value calculation in annotation is tricky, simplified here.
        # Ideally handled in detail view or subquery for accuracy.
        total_return_value=Sum(
            F('salesreturn__items__quantity') * F('salesreturn__items__unit_price'),
            distinct=True
        ),
        total_refund_paid=Sum(
            'refunds__refund_amount', 
            filter=Q(refunds__status='Paid'),
            distinct=True
        ),
        total_order_count=Count('salesorder', distinct=True),
        cancelled_order_count=Count(
            'salesorder', 
            filter=Q(salesorder__status='cancelled'), 
            distinct=True
        ),
        last_order_date=Max('salesorder__created_at')
    )

    # Export Logic
    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="customers_detailed.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'Name', 'Phone', 'Total Orders', 'Total Sales', 
            'Total Returns', 'Due Amount', 'Return Rate %', 
            'Cancelled Orders', 'Last Order Date'
        ])
        for c in customers_list:
            t_sales = c.total_sales_amount or 0
            t_returns = c.total_return_value or 0
            # Fallback for returns if annotation misses
            if t_returns == 0:
                 val = SalesReturn.objects.filter(customer=c).aggregate(
                    total=Sum(F('items__quantity') * F('items__unit_price'))
                 )['total']
                 t_returns = val or 0

            t_received = c.total_received_amount or 0
            t_refunds = c.total_refund_paid or 0
            current_due = (t_sales - t_returns) - (t_received - t_refunds)
            
            return_rate = 0
            if t_sales > 0:
                return_rate = (t_returns / t_sales) * 100

            last_order = c.last_order_date.strftime("%Y-%m-%d") if c.last_order_date else "N/A"

            writer.writerow([
                c.name, c.phone, c.total_order_count, round(t_sales, 2), 
                round(t_returns, 2), round(current_due, 2),
                f"{return_rate:.1f}%", c.cancelled_order_count, last_order
            ])
        return response

    if request.GET.get("export") == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Analysis"
        headers = [
            'Name', 'Email', 'Phone', 'Address', 
            'Total Orders', 'Total Sales', 'Total Returns', 
            'Due Amount', 'Return Rate %', 'Cancelled', 'Last Order Date'
        ]
        ws.append(headers)
        for c in customers_list:
            t_sales = c.total_sales_amount or 0
            t_returns = c.total_return_value or 0
            if t_returns == 0:
                 val = SalesReturn.objects.filter(customer=c).aggregate(
                    total=Sum(F('items__quantity') * F('items__unit_price'))
                 )['total']
                 t_returns = val or 0
            
            t_received = c.total_received_amount or 0
            t_refunds = c.total_refund_paid or 0
            current_due = (t_sales - t_returns) - (t_received - t_refunds)
            
            return_rate = 0
            if t_sales > 0:
                return_rate = (t_returns / t_sales) * 100

            last_order = c.last_order_date.strftime("%d-%b-%Y") if c.last_order_date else "-"

            ws.append([
                c.name, c.email, c.phone, c.address,
                c.total_order_count, t_sales, t_returns,
                current_due, f"{return_rate:.2f}%",
                c.cancelled_order_count, last_order
            ])
        # Auto-adjust columns
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customer_analysis.xlsx"'
        wb.save(response)
        return response

    paginator = Paginator(customers_list, 15)
    page_number = request.GET.get('page')
    customers = paginator.get_page(page_number)

    # Calculation for display
    for customer in customers:
        t_sales = customer.total_sales_amount or 0
        t_received = customer.total_received_amount or 0
        t_returns = customer.total_return_value or 0
        t_refunds = customer.total_refund_paid or 0

        # Precision Fix for Returns if 0
        if t_returns == 0:
            val = SalesReturn.objects.filter(customer=customer).aggregate(
                total=Sum(F('items__quantity') * F('items__unit_price'))
            )['total']
            t_returns = val or 0

        net_sales = t_sales - t_returns
        net_payment = t_received - t_refunds
        current_due = net_sales - net_payment
        
        customer.t_sales = t_sales
        customer.t_received = t_received
        customer.t_returns = t_returns
        customer.t_refunds = t_refunds
        customer.current_due = current_due
        customer.order_count = customer.total_order_count
        customer.last_order = customer.last_order_date

    context = {
        'customers': customers,
        'title': 'All Customers',
        'filter_form': filter_form
    }
    return render(request, 'partners/customer_list.html', context)

@login_required
@permission_required('partners.add_customer', login_url='/admin/')
def add_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Customer added successfully!')
            return redirect('partners:customer_list') 
    else:
        form = CustomerForm()
    return render(request, 'partners/add_customer.html', {'form': form, 'title': 'Add New Customer'})

@login_required
@permission_required('partners.change_customer', login_url='/admin/')
def edit_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f'Customer "{customer.name}" updated successfully!')
            return redirect('partners:customer_list')
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'partners/edit_customer.html', {'form': form, 'title': f'Edit Customer: {customer.name}'})

@login_required
@permission_required('partners.delete_customer', login_url='/admin/')
def delete_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer_name = customer.name
        customer.delete()
        messages.success(request, f'Customer "{customer_name}" deleted successfully!')
        return redirect('partners:customer_list')
    return render(request, 'confirm_delete.html', {'object': customer, 'title': f'Confirm Delete: {customer.name}'})


# --- AJAX Views ---
@login_required
@permission_required('partners.add_supplier', raise_exception=True)
def ajax_add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()
            return JsonResponse({'success': True, 'id': supplier.id, 'name': supplier.name})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)

@login_required
@permission_required('partners.add_customer', raise_exception=True)
def ajax_add_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save()
            return JsonResponse({'success': True, 'id': customer.id, 'name': customer.name})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)

@login_required
@permission_required('partners.view_supplier', login_url='/admin/')
def supplier_detail(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)

    all_bills_qs = VendorBill.objects.filter(supplier=supplier).exclude(status='Cancelled')
    total_bill_amount = all_bills_qs.aggregate(total=Coalesce(Sum('total_amount'), Decimal(0)))['total']

    all_payments_qs = BillPayment.objects.filter(vendor_bill__supplier=supplier)
    payments_agg = all_payments_qs.aggregate(
        paid=Coalesce(Sum('amount_paid'), Decimal(0)),
        credit=Coalesce(Sum('credit_applied'), Decimal(0))
    )
    total_paid_amount = payments_agg['paid'] + payments_agg['credit']
    all_shipped_returns_qs = PurchaseReturn.objects.filter(supplier=supplier, status__in=['shipped', 'completed'])
    
    total_return_value = Decimal(0)
    if all_shipped_returns_qs.exists():
        total_return_value = all_shipped_returns_qs.aggregate(
            total=Coalesce(Sum(F('items__quantity') * F('items__unit_price')), Decimal(0))
        )['total']

    available_credit_notes = SupplierCreditNote.objects.filter(supplier=supplier, status='Available', amount_available__gt=0)
    total_available_credit = available_credit_notes.aggregate(total=Coalesce(Sum('amount_available'), Decimal(0)))['total']
    
    current_balance = total_bill_amount - total_paid_amount - total_available_credit

    context = {
        'supplier': supplier,
        'all_bills': all_bills_qs.order_by('-bill_date'),
        'all_payments': all_payments_qs.order_by('-payment_date'),
        'all_returns': all_shipped_returns_qs.order_by('-return_date'),
        'available_credit_notes': available_credit_notes.order_by('created_date'),
        'total_available_credit': total_available_credit,
        'total_bill_amount': total_bill_amount,
        'total_paid_amount': total_paid_amount,
        'total_return_amount': total_return_value,
        'current_balance': current_balance,
        'title': f'Supplier Details: {supplier.name}'
    }
    return render(request, 'partners/supplier_detail.html', context)

@login_required
@permission_required('partners.view_customer', login_url='/admin/')
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)

    # ১. গিফট কার্ড সেকশন (Optimized)
    gift_cards = GiftCard.objects.filter(customer=customer).order_by('-id')
    total_card_balance = GiftCard.objects.filter(
        customer=customer, 
        is_active=True
    ).aggregate(total=Sum('current_balance'))['total'] or 0

    # ২. ইনভয়েস এবং পেমেন্ট সেকশন
    # [NOTE] CustomerInvoice এখন finance.ar.models থেকে সঠিকভাবে ইম্পোর্ট হচ্ছে
    all_invoices = CustomerInvoice.objects.filter(customer=customer).select_related('sales_order').order_by('-invoice_date')
    total_invoiced_amount = all_invoices.aggregate(total=Sum('total_amount'))['total'] or 0

    all_payments = InvoicePayment.objects.filter(customer_invoice__customer=customer).order_by('-payment_date')
    total_paid_amount = all_payments.aggregate(total=Sum('amount_received'))['total'] or 0

    # ৩. রিটার্ন সেকশন (Corrected with SalesReturnItem)
    # [NOTE] SalesReturnItem এখন sales.models থেকে সঠিকভাবে ইম্পোর্ট হচ্ছে
    total_return_amount = SalesReturnItem.objects.filter(
        sales_return__customer=customer
    ).aggregate(
        grand_total=Sum(F('quantity') * F('unit_price'))
    )['grand_total'] or 0

    all_returns = SalesReturn.objects.filter(customer=customer).annotate(
        total_amount=Sum(F('items__quantity') * F('items__unit_price'))
    ).order_by('-return_date')

    # ৪. POS সেলস
    all_pos_orders = POSOrder.objects.filter(customer=customer).order_by('-order_date')
    total_pos_sales = all_pos_orders.aggregate(total=Sum('net_amount'))['total'] or 0

    # ৫. ডিউ ক্যালকুলেশন
    current_due = total_invoiced_amount - total_paid_amount - total_return_amount

    context = {
        'customer': customer,
        'gift_cards': gift_cards,
        'total_card_balance': total_card_balance,
        'all_invoices': all_invoices,
        'all_payments': all_payments,
        'all_returns': all_returns,
        'all_pos_orders': all_pos_orders,
        'total_invoiced_amount': total_invoiced_amount,
        'total_paid_amount': total_paid_amount,
        'total_return_amount': total_return_amount,
        'total_pos_sales': total_pos_sales,
        'current_due': current_due,
        'title': f'Customer Details: {customer.name}'
    }
    return render(request, 'partners/customer_detail.html', context)